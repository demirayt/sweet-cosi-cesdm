"""cesdm.domain.model.excel — Excel persistence

Spreadsheet import/export using the CESDM asset/view sheet
abbreviation convention.

Auto-extracted from the legacy monolithic module as part of the
package-hierarchy refactor (see docs/architecture/package_layout.md).
Behaviour is unchanged; only module boundaries moved.
"""

from __future__ import annotations

from typing import Any, ClassVar, Dict, List, Optional, Union
import os
import pathlib
import re
import yaml


class ExcelMixin:
    """Mixin — see module docstring for the responsibility this covers."""

    def import_excel_flat(
        self,
        path: Union[str, pathlib.Path],
        *,
        skip_unknown_classes: bool = True,
        skip_unknown_fields:  bool = True,
    ) -> Dict[str, int]:
        """
        Import a flat Excel workbook produced by :meth:`export_excel_flat`.

        One sheet per entity class, one row per entity, one column per
        attribute or relation.  The first column must be ``entity_id``.

        The schema is used to route each column to
        ``add_attribute`` or ``add_relation`` automatically.

        Parameters
        ----------
        path :
            Path to the flat Excel workbook (.xlsx).
        skip_unknown_classes :
            If True (default), silently skip sheets whose name does not
            match any schema class.  If False, raise ``KeyError``.
        skip_unknown_fields :
            If True (default), silently skip columns not declared in the
            schema.  If False, raise ``KeyError``.

        Returns
        -------
        dict
            ``{class_name: n_entities_imported}``
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "openpyxl is required.  pip install openpyxl"
            )

        p = pathlib.Path(path)
        wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
        stats: Dict[str, int] = {}

        for ws in wb.worksheets:
            cname = ws.title
            cdef  = self.classes.get(cname)
            if cdef is None:
                if not skip_unknown_classes:
                    raise KeyError(f"Sheet '{cname}' is not a known class")
                continue

            attrs_def, rels_def = self._collect_inherited_fields(cdef)
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            header = [str(c).strip() if c is not None else "" for c in rows[0]]

            count = 0
            for row in rows[1:]:
                if not any(v is not None for v in row):
                    continue
                row_dict = dict(zip(header, row))
                eid = str(row_dict.get("entity_id", "")).strip()
                if not eid:
                    continue

                if eid not in self.entities.get(cname, {}):
                    self.add_entity(cname, eid)

                for col, val in row_dict.items():
                    if col == "entity_id" or val is None or str(val).strip() == "":
                        continue
                    if col in rels_def:
                        # Relation — value may be a single id or JSON array
                        val_str = str(val).strip()
                        if val_str.startswith("["):
                            import json as _j
                            try:
                                targets = _j.loads(val_str)
                                for t in targets:
                                    if t:
                                        self.add_relation(eid, col, str(t))
                            except Exception:
                                if val_str:
                                    self.add_relation(eid, col, val_str)
                        else:
                            self.add_relation(eid, col, val_str)
                    elif col in attrs_def:
                        self.add_attribute(eid, col, val)
                    else:
                        if not skip_unknown_fields:
                            raise KeyError(
                                f"Column '{col}' in sheet '{cname}' is not "
                                f"a known attribute or relation"
                            )
                count += 1
            stats[cname] = count

        wb.close()
        return stats

    # ================================================================== #
    #  import_csv_by_class_wide                                            #
    # ================================================================== #

    def export_excel(
        self,
        path: Union[str, pathlib.Path],
        *,
        include_non_assets: bool = True,
    ) -> None:
        """
        Export the model to an Excel workbook (.xlsx) where each sheet
        represents one entity class or one representation view.

        Sheet layout
        ------------
        **Non-asset sheets** (EnergyCarrier, NetworkNode, GeographicalRegion,
        GeneratorType, StorageType, …):
          One sheet per class, named after the class.  Columns: ``entity_id``
          followed by all attribute and relation fields.  These are flat — no
          view nesting.

        **Asset identity sheets** (GenerationUnit, StorageUnit, …):
          One sheet per asset class containing only the identity columns
          (``asset_id``, own attributes, ``hasTechnology``).  Sheet name is
          the class name.

        **View sheets** (SinglePort.TopologyView, Generation.DispatchView, …):
          One sheet per view class that is actually populated, named
          ``<AssetClass>.<ViewClass>`` (e.g. ``GenerationUnit.Generation.DispatchView``).
          Columns: ``asset_id`` (the owner asset) followed by all view
          attribute and relation fields (``representsAsset`` omitted — it is
          implicit in the ``asset_id`` column and the sheet name).

        All sheets share the same ``asset_id`` / ``entity_id`` primary key so
        they can be joined in Excel via VLOOKUP or Power Query.

        Formatting
        ----------
        - Header row: bold, light-blue fill, auto-filtered.
        - Column widths auto-sized (capped at 50).
        - Alternating row fill for readability.
        - Font: Calibri 11.

        Parameters
        ----------
        path :
            Output file path.  Parent directories are created if absent.
            Extension should be ``.xlsx``.
        include_non_assets :
            If False, only asset and view sheets are written (useful when
            non-asset data is managed elsewhere).
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import (Font, PatternFill, Alignment,
                                          Border, Side)
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install it with: pip install openpyxl"
            )

        p = pathlib.Path(path)
        p.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        wb.remove(wb.active)  # remove default empty sheet

        # ── style constants ───────────────────────────────────────────────
        HDR_FONT  = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        HDR_FILL  = PatternFill("solid", start_color="1F4E79")   # dark blue
        ALT_FILL  = PatternFill("solid", start_color="DCE6F1")   # light blue
        STD_FONT  = Font(name="Calibri", size=11)
        THIN      = Side(style="thin", color="B8CCE4")
        HDR_BORDER = Border(bottom=Side(style="medium", color="1F4E79"))

        def _make_sheet(title: str) -> "Worksheet":
            ws = wb.create_sheet(title=title[:31])  # Excel limit 31 chars
            return ws

        def _write_header(ws, cols: List[str]) -> None:
            for ci, col in enumerate(cols, 1):
                cell = ws.cell(row=1, column=ci, value=col)
                cell.font   = HDR_FONT
                cell.fill   = HDR_FILL
                cell.alignment = Alignment(horizontal="center", vertical="center",
                                           wrap_text=False)
                cell.border = HDR_BORDER
            ws.row_dimensions[1].height = 18
            ws.auto_filter.ref = ws.dimensions

        def _write_rows(ws, cols: List[str], rows: List[Dict]) -> None:
            for ri, row in enumerate(rows, 2):
                fill = ALT_FILL if ri % 2 == 0 else None
                for ci, col in enumerate(cols, 1):
                    val = row.get(col, "")
                    cell = ws.cell(row=ri, column=ci, value=val if val != "" else None)
                    cell.font = STD_FONT
                    if fill:
                        cell.fill = fill

        def _autosize(ws, cols: List[str]) -> None:
            for ci, col in enumerate(cols, 1):
                letter = get_column_letter(ci)
                max_w = max(
                    len(str(col)),
                    *(len(str(ws.cell(r, ci).value or ""))
                      for r in range(1, ws.max_row + 1))
                )
                ws.column_dimensions[letter].width = min(max_w + 2, 50)

        # ── shared helpers ────────────────────────────────────────────────
        def _unwrap(raw) -> str:
            if isinstance(raw, dict):
                v = raw.get("value", "")
                u = raw.get("unit", "")
                if u:
                    return f"{'' if v is None else v} [{u}]"
                return "" if v is None else str(v)
            return "" if raw is None else str(raw)

        def _entity_row(ent, cname: str,
                        id_col: str, id_val: str,
                        skip_rels: Optional[List[str]] = None) -> Dict[str, str]:
            cdef = self.classes.get(cname)
            if not cdef:
                return {id_col: id_val}
            attrs_def, refs_def = self._collect_inherited_fields(cdef)
            data = getattr(ent, "data", {}) or {}
            skip = set(skip_rels or [])
            row: Dict[str, str] = {id_col: id_val}
            for rname in refs_def:
                if rname in skip or rname not in data or data[rname] in ("", None):
                    continue
                val = data[rname]
                targets = [str(x) for x in val if x not in ("", None)] \
                    if isinstance(val, (list, tuple)) else [str(val)]
                row[rname] = " | ".join(targets)
            for aname in attrs_def:
                if aname in data and data[aname] not in ("", None):
                    row[aname] = _unwrap(data[aname])
            return row

        def _cols_for(cname: str,
                      id_col: str,
                      skip_rels: Optional[List[str]] = None) -> List[str]:
            cdef = self.classes.get(cname)
            if not cdef:
                return [id_col]
            attrs_def, refs_def = self._collect_inherited_fields(cdef)
            skip = set(skip_rels or [])
            cols = [id_col]
            for name in refs_def:
                if name not in skip:
                    cols.append(name)
            cols.extend(attrs_def.keys())
            return cols

        view_index = self._build_view_index()

        # ── non-asset sheets ──────────────────────────────────────────────
        if include_non_assets:
            for ncls in sorted(self._discover_non_asset_classes()):
                ents = self.entities.get(ncls) or {}
                if not ents:
                    continue
                cols = _cols_for(ncls, "entity_id")
                rows = [_entity_row(ent, ncls, "entity_id", eid)
                        for eid, ent in ents.items()]
                ws = _make_sheet(ncls[:31])
                _write_header(ws, cols)
                _write_rows(ws, cols, rows)
                _autosize(ws, cols)

        # ── asset identity sheets + view sheets ───────────────────────────
        for acls in sorted(self._discover_asset_classes()):
            ents = self.entities.get(acls) or {}
            if not ents:
                continue

            allowed_views = self._discover_view_map().get(acls, [])

            # Identity sheet
            id_cols = _cols_for(acls, "asset_id")
            id_rows = [_entity_row(ent, acls, "asset_id", eid)
                       for eid, ent in ents.items()]
            ws = _make_sheet(acls[:31])
            _write_header(ws, id_cols)
            _write_rows(ws, id_cols, id_rows)
            _autosize(ws, id_cols)

            # One sheet per allowed view class that has data
            for vcls in allowed_views:
                vcls_ents_for_asset = {
                    eid: view_index[eid][vcls]
                    for eid in ents
                    if eid in view_index and vcls in view_index[eid]
                }
                if not vcls_ents_for_asset:
                    continue

                vcols = _cols_for(vcls, "asset_id",
                                   skip_rels=[self._REPRESENTS_ASSET_REL])
                vrows = [
                    _entity_row(vent, vcls, "asset_id", eid,
                                skip_rels=[self._REPRESENTS_ASSET_REL])
                    for eid, vent in vcls_ents_for_asset.items()
                ]
                # Sheet name: "<AssetAbbr>.<ViewAbbr>" — both fit in 31 chars
                # and are reversible on import via _abbrev_for().
                a_abbr = self._abbrev_for(acls)
                v_abbr = self._abbrev_for(vcls)
                sheet_name = f"{a_abbr}.{v_abbr}"
                ws = _make_sheet(sheet_name)
                _write_header(ws, vcols)
                _write_rows(ws, vcols, vrows)
                _autosize(ws, vcols)

        wb.save(str(p))

    def import_excel(
        self,
        path: Union[str, pathlib.Path],
        *,
        strict_unknown: bool = False,
    ) -> Dict[str, Any]:
        """
        Import entities from an Excel workbook produced by
        :meth:`export_excel`.

        Sheet recognition
        -----------------
        Sheets are classified by name:

        - **Non-asset sheets**: name matches a domain-role class
          (derived from schema inheritance). ``entity_id`` is the primary key.
        - **Asset identity sheets**: name matches an asset-role class.
          ``asset_id`` column is the primary key.
        - **View sheets**: name has the form ``<AssetClass>.<ViewClass>``.
          ``asset_id`` column links to the owner asset; ``representsAsset``
          is injected automatically.
        - **Unknown sheet names** are skipped with a warning in the summary.

        Values stored as ``"value [unit]"`` strings (from export) are split
        back into separate value and unit fields on import.  Plain scalars
        and pipe-separated relation lists are handled transparently.

        Parameters
        ----------
        path :
            Input ``.xlsx`` file path.
        strict_unknown :
            If True, unknown column names are added to the unknowns list.

        Returns
        -------
        dict
            Summary with keys ``created_entities``, ``set_attributes``,
            ``set_relations``, ``unknowns``, ``skipped_sheets``.
        """
        try:
            import openpyxl as _opxl
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel import. "
                "Install it with: pip install openpyxl"
            )

        p = pathlib.Path(path)
        wb = _opxl.load_workbook(str(p), read_only=True, data_only=True)

        created = set_attr = set_ref = 0
        unknowns: List[tuple] = []
        skipped: List[str] = []

        asset_cls_set     = self._discover_asset_classes()
        non_asset_cls_set = self._discover_non_asset_classes()

        # Reverse abbreviation maps for import
        # Reverse abbrev maps derived from schema classes
        _abbr_to_asset = {self._abbrev_for(c): c
                          for c in self._discover_asset_classes()}
        _abbr_to_view  = {self._abbrev_for(c): c
                          for c in self._discover_view_classes()}

        # Pre-collect known fields per class
        _known_attrs: Dict[str, set] = {}
        _known_refs:  Dict[str, set] = {}
        _refs_def_map: Dict[str, Any] = {}
        for cname, cdef in self.classes.items():
            a, r = self._collect_inherited_fields(cdef)
            _known_attrs[cname] = set(a.keys())
            _known_refs[cname]  = set(r.keys())
            _refs_def_map[cname] = r

        def _ensure(cls: str, eid: str) -> None:
            nonlocal created
            if cls not in self.entities:
                self.entities[cls] = {}
            if eid not in self.entities[cls]:
                self.add_entity(cls, eid)
                created += 1

        def _parse_val(raw_str: str):
            """Split 'value [unit]' → (value_str, unit_str) or (raw, None)."""
            s = str(raw_str).strip()
            if s.endswith("]") and " [" in s:
                val_part, unit_part = s.rsplit(" [", 1)
                return val_part.strip(), unit_part.rstrip("]").strip()
            return s, None

        def _set_attr(cls: str, eid: str, aname: str, raw) -> None:
            nonlocal set_attr
            if aname not in _known_attrs.get(cls, set()):
                if strict_unknown:
                    unknowns.append((cls, eid, f"unknown attribute: {aname}"))
                return
            if raw is None or str(raw).strip() == "":
                return
            val_str, unit_str = _parse_val(raw)
            if unit_str:
                coerced = self._coerce_for_attr(cls, aname, val_str)
                self.add_attribute(eid, aname, {"value": coerced, "unit": unit_str})
            else:
                coerced = self._coerce_for_attr(cls, aname, val_str)
                self.add_attribute(eid, aname, coerced)
            set_attr += 1

        def _set_rel(cls: str, eid: str, rname: str, raw) -> None:
            nonlocal set_ref
            if rname not in _known_refs.get(cls, set()):
                if strict_unknown:
                    unknowns.append((cls, eid, f"unknown relation: {rname}"))
                return
            if raw is None or str(raw).strip() == "":
                return
            txt = str(raw).strip()
            targets = [t.strip() for t in txt.split("|") if t.strip()]
            for tgt in targets:
                self.add_relation(eid, rname, tgt)
                set_ref += 1

        def _ingest_sheet(ws, cls: str, id_col: str,
                          skip_rels: Optional[set] = None,
                          inject_rel: Optional[tuple] = None) -> None:
            """Read one sheet into cls entities."""
            skip = skip_rels or set()
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return
            header = [str(c).strip() if c is not None else "" for c in rows[0]]
            if id_col not in header:
                unknowns.append((cls, None,
                    f"Sheet '{ws.title}' missing '{id_col}' column"))
                return
            id_idx = header.index(id_col)
            known_a = _known_attrs.get(cls, set())
            known_r = _known_refs.get(cls, set())

            for row in rows[1:]:
                if not any(c is not None for c in row):
                    continue
                eid = str(row[id_idx]).strip() if row[id_idx] is not None else ""
                if not eid:
                    continue
                _ensure(cls, eid)
                if inject_rel:
                    rel_name, rel_target = inject_rel
                    if rel_name in _known_refs.get(cls, set()):
                        self.add_relation(eid, rel_name, rel_target)

                for ci, col in enumerate(header):
                    if ci == id_idx or not col or col in skip:
                        continue
                    raw = row[ci]
                    if raw is None or str(raw).strip() == "":
                        continue
                    if col in known_r:
                        _set_rel(cls, eid, col, raw)
                    elif col in known_a:
                        _set_attr(cls, eid, col, raw)
                    elif strict_unknown:
                        unknowns.append((cls, eid, f"unknown column: {col}"))

        def _view_id(vcls: str, asset_id: str) -> str:
            snake = re.sub(r"(?<!^)(?=[A-Z])", "_", vcls).lower()
            return f"{snake}.{asset_id}"

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Non-asset sheet
            if sheet_name in non_asset_cls_set:
                _ingest_sheet(ws, sheet_name, "entity_id")
                continue

            # Asset identity sheet
            if sheet_name in asset_cls_set:
                _ingest_sheet(ws, sheet_name, "asset_id")
                continue

            # View sheet: "AssetAbbr.ViewAbbr" (from export_excel abbreviations)
            if "." in sheet_name:
                parts = sheet_name.split(".", 1)
                a_abbr, v_abbr = parts[0], parts[1]
                acls = _abbr_to_asset.get(a_abbr, a_abbr)
                vcls = _abbr_to_view.get(v_abbr, v_abbr)
                if acls in asset_cls_set and vcls in self.classes:
                    # Read rows: asset_id → create view entity, inject representsAsset
                    rows = list(ws.iter_rows(values_only=True))
                    if not rows:
                        continue
                    header = [str(c).strip() if c is not None else ""
                              for c in rows[0]]
                    if "asset_id" not in header:
                        unknowns.append((vcls, None,
                            f"Sheet '{sheet_name}' missing 'asset_id'"))
                        continue
                    aid_idx = header.index("asset_id")
                    known_a = _known_attrs.get(vcls, set())
                    known_r = _known_refs.get(vcls, set())
                    ra = self._REPRESENTS_ASSET_REL

                    for row in rows[1:]:
                        if not any(c is not None for c in row):
                            continue
                        asset_id = (str(row[aid_idx]).strip()
                                    if row[aid_idx] is not None else "")
                        if not asset_id:
                            continue
                        vid = _view_id(vcls, asset_id)
                        _ensure(vcls, vid)
                        if ra in known_r:
                            self.add_relation(vid, ra, asset_id)
                            set_ref += 1

                        for ci, col in enumerate(header):
                            if ci == aid_idx or not col or col == ra:
                                continue
                            raw = row[ci]
                            if raw is None or str(raw).strip() == "":
                                continue
                            if col in known_r:
                                _set_rel(vcls, vid, col, raw)
                            elif col in known_a:
                                _set_attr(vcls, vid, col, raw)
                            elif strict_unknown:
                                unknowns.append((vcls, vid,
                                    f"unknown column: {col}"))
                    continue

            skipped.append(sheet_name)

        wb.close()
        return {
            "created_entities": created,
            "set_attributes":   set_attr,
            "set_relations":    set_ref,
            "unknowns":         unknowns,
            "skipped_sheets":   skipped,
        }

    def export_excel_flat(
        self,
        path: Union[str, pathlib.Path],
        *,
        include_non_assets: bool = True,
    ) -> None:
        """
        Export the model to a flat Excel workbook (.xlsx).

        One sheet per entity class, one row per entity, one column per
        attribute or relation — the same layout as the wide CSV export
        but in a single Excel file.

        Unlike :meth:`export_excel` (which splits asset identity and views
        across separate sheets), every field of a class appears on one sheet
        so each row is fully self-contained.

        Parameters
        ----------
        path :
            Output file path (.xlsx). Parent directories are created if absent.
        include_non_assets :
            If False, skip non-asset / non-view classes (carriers, regions …).
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install it with: pip install openpyxl"
            )
        import json as _json

        p = pathlib.Path(path)
        p.parent.mkdir(parents=True, exist_ok=True)

        hdr_font   = Font(name="Calibri", size=11, bold=True)
        hdr_fill   = PatternFill("solid", fgColor="BDD7EE")
        row_fill_a = PatternFill("solid", fgColor="FFFFFF")
        row_fill_b = PatternFill("solid", fgColor="F2F2F2")
        body_font  = Font(name="Calibri", size=11)
        top_align  = Alignment(wrap_text=False, vertical="top")

        def _unwrap(raw) -> str:
            if isinstance(raw, dict):
                v = raw.get("value", "")
                return "" if v is None else str(v)
            return "" if raw is None else str(raw)

        def _write_sheet(wb, sheet_name, header, rows):
            ws = wb.create_sheet(title=sheet_name[:31])
            for ci, col in enumerate(header, 1):
                cell = ws.cell(row=1, column=ci, value=col)
                cell.font      = hdr_font
                cell.fill      = hdr_fill
                cell.alignment = top_align
            ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}1"

            for ri, row in enumerate(rows, 2):
                fill = row_fill_a if ri % 2 == 0 else row_fill_b
                for ci, col in enumerate(header, 1):
                    cell = ws.cell(row=ri, column=ci, value=row.get(col, ""))
                    cell.font      = body_font
                    cell.fill      = fill
                    cell.alignment = top_align

            for ci, col in enumerate(header, 1):
                max_len = max(
                    (len(str(row.get(col, "") or "")) for row in rows),
                    default=0,
                )
                ws.column_dimensions[get_column_letter(ci)].width = (
                    min(max(max_len, len(col)) + 2, 50)
                )

        wb = Workbook()
        wb.remove(wb.active)
        non_asset_cls = self._discover_non_asset_classes()

        for cname, cdef in self.classes.items():
            ents = self.entities.get(cname, {})
            if not ents:
                continue
            if not include_non_assets and cname in non_asset_cls:
                continue

            attrs_def, refs_def = self._collect_inherited_fields(cdef)
            header = ["entity_id"] + list(refs_def.keys()) + list(attrs_def.keys())
            rows   = []

            for eid, ent in ents.items():
                row  = {h: "" for h in header}
                row["entity_id"] = eid
                data = getattr(ent, "data", {}) or {}

                for rn in refs_def:
                    val = data.get(rn)
                    if val not in (None, ""):
                        row[rn] = (
                            _json.dumps(val, ensure_ascii=False)
                            if isinstance(val, (list, dict, tuple)) else str(val)
                        )
                for an in attrs_def:
                    val = data.get(an)
                    if val not in (None, ""):
                        row[an] = _unwrap(val)

                rows.append(row)

            _write_sheet(wb, cname, header, rows)

        wb.save(str(p))
